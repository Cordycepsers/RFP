"""
Excel tracker generator for Proposaland opportunity monitoring.
Creates professional Excel trackers with reference number columns and comprehensive data.
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path
from loguru import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelGenerator:
    """Professional Excel tracker generator with formatting and reference numbers."""
    
    def __init__(self, config: Dict):
        self.config = config
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
        
    def generate_tracker(self, opportunities: List[Dict], filename: str = None) -> str:
        """Generate comprehensive Excel tracker with multiple sheets."""
        if not filename:
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f"proposaland_tracker_{date_str}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            # Create workbook with multiple sheets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Main opportunities sheet
                self._create_opportunities_sheet(opportunities, writer)
                
                # Summary sheet
                self._create_summary_sheet(opportunities, writer)
                
                # Reference numbers sheet
                self._create_reference_sheet(opportunities, writer)
                
                # Priority breakdown sheet
                self._create_priority_sheet(opportunities, writer)
            
            # Apply formatting
            self._apply_formatting(filepath)
            
            logger.info(f"Generated Excel tracker: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating Excel tracker: {e}")
            raise
    
    def _create_opportunities_sheet(self, opportunities: List[Dict], writer):
        """Create main opportunities sheet with all details."""
        # Prepare data for DataFrame
        rows = []
        
        for i, opp in enumerate(opportunities, 1):
            row = {
                'ID': i,
                'Title': opp.get('title', ''),
                'Organization': opp.get('organization', ''),
                'Reference Number': opp.get('reference_number', ''),
                'Reference Confidence': opp.get('reference_confidence', 0.0),
                'Priority': opp.get('priority', 'Low'),
                'Relevance Score': opp.get('relevance_score', 0.0),
                'Keywords Found': ', '.join(opp.get('keywords_found', [])),
                'Budget (USD)': opp.get('budget', ''),
                'Currency': opp.get('currency', 'USD'),
                'Location': opp.get('location', ''),
                'Deadline': self._format_deadline(opp.get('deadline')),
                'Days Until Deadline': self._calculate_days_until_deadline(opp.get('deadline')),
                'Source URL': opp.get('source_url', ''),
                'Description': opp.get('description', '')[:500] + ('...' if len(opp.get('description', '')) > 500 else ''),
                'Extracted Date': self._format_date(opp.get('extracted_date')),
                'Status': 'New',
                'Notes': '',
                'Action Required': self._determine_action(opp),
                'Follow-up Date': '',
                'Contact Person': '',
                'Proposal Submitted': 'No',
                'Outcome': ''
            }
            rows.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(rows)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Opportunities', index=False)
    
    def _create_summary_sheet(self, opportunities: List[Dict], writer):
        """Create summary statistics sheet."""
        total_count = len(opportunities)
        
        # Priority breakdown
        priority_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
        for opp in opportunities:
            priority = opp.get('priority', 'Low')
            priority_counts[priority] += 1
        
        # Organization breakdown
        org_counts = {}
        for opp in opportunities:
            org = opp.get('organization', 'Unknown')
            org_counts[org] = org_counts.get(org, 0) + 1
        
        # Keyword breakdown
        keyword_counts = {}
        for opp in opportunities:
            for keyword in opp.get('keywords_found', []):
                keyword_counts[keyword] = keyword_counts.get(keyword, 0) + 1
        
        # Budget analysis
        budgets = [opp.get('budget') for opp in opportunities if opp.get('budget')]
        budget_stats = {
            'Total Opportunities with Budget': len(budgets),
            'Average Budget': sum(budgets) / len(budgets) if budgets else 0,
            'Min Budget': min(budgets) if budgets else 0,
            'Max Budget': max(budgets) if budgets else 0
        }
        
        # Create summary data
        summary_data = []
        
        # General statistics
        summary_data.extend([
            ['GENERAL STATISTICS', ''],
            ['Total Opportunities', total_count],
            ['Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            
            ['PRIORITY BREAKDOWN', ''],
            ['Critical Priority', priority_counts['Critical']],
            ['High Priority', priority_counts['High']],
            ['Medium Priority', priority_counts['Medium']],
            ['Low Priority', priority_counts['Low']],
            ['', ''],
            
            ['BUDGET ANALYSIS', ''],
            ['Opportunities with Budget Info', budget_stats['Total Opportunities with Budget']],
            ['Average Budget (USD)', f"${budget_stats['Average Budget']:,.2f}" if budget_stats['Average Budget'] else 'N/A'],
            ['Minimum Budget (USD)', f"${budget_stats['Min Budget']:,.2f}" if budget_stats['Min Budget'] else 'N/A'],
            ['Maximum Budget (USD)', f"${budget_stats['Max Budget']:,.2f}" if budget_stats['Max Budget'] else 'N/A'],
            ['', '']
        ])
        
        # Top organizations
        top_orgs = sorted(org_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        summary_data.append(['TOP ORGANIZATIONS', ''])
        for org, count in top_orgs:
            summary_data.append([org, count])
        
        summary_data.append(['', ''])
        
        # Top keywords
        top_keywords = sorted(keyword_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        summary_data.append(['TOP KEYWORDS', ''])
        for keyword, count in top_keywords:
            summary_data.append([keyword, count])
        
        # Create DataFrame
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_reference_sheet(self, opportunities: List[Dict], writer):
        """Create reference numbers analysis sheet."""
        ref_data = []
        
        for i, opp in enumerate(opportunities, 1):
            ref_number = opp.get('reference_number', '')
            if ref_number:
                ref_data.append({
                    'ID': i,
                    'Title': opp.get('title', ''),
                    'Organization': opp.get('organization', ''),
                    'Reference Number': ref_number,
                    'Confidence Score': opp.get('reference_confidence', 0.0),
                    'Pattern Type': opp.get('scoring_details', {}).get('reference_pattern', 'Unknown'),
                    'Organization Type': opp.get('scoring_details', {}).get('org_type', 'Unknown'),
                    'Priority': opp.get('priority', 'Low'),
                    'Source URL': opp.get('source_url', '')
                })
        
        if ref_data:
            ref_df = pd.DataFrame(ref_data)
            ref_df = ref_df.sort_values('Confidence Score', ascending=False)
            ref_df.to_excel(writer, sheet_name='Reference Numbers', index=False)
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=[
                'ID', 'Title', 'Organization', 'Reference Number', 
                'Confidence Score', 'Pattern Type', 'Organization Type', 
                'Priority', 'Source URL'
            ])
            empty_df.to_excel(writer, sheet_name='Reference Numbers', index=False)
    
    def _create_priority_sheet(self, opportunities: List[Dict], writer):
        """Create priority-based breakdown sheet."""
        priorities = ['Critical', 'High', 'Medium', 'Low']
        
        for priority in priorities:
            priority_opps = [opp for opp in opportunities if opp.get('priority') == priority]
            
            if priority_opps:
                priority_data = []
                for i, opp in enumerate(priority_opps, 1):
                    priority_data.append({
                        'Rank': i,
                        'Title': opp.get('title', ''),
                        'Organization': opp.get('organization', ''),
                        'Reference Number': opp.get('reference_number', ''),
                        'Score': opp.get('relevance_score', 0.0),
                        'Keywords': ', '.join(opp.get('keywords_found', [])),
                        'Deadline': self._format_deadline(opp.get('deadline')),
                        'Budget': opp.get('budget', ''),
                        'Source URL': opp.get('source_url', '')
                    })
                
                priority_df = pd.DataFrame(priority_data)
                sheet_name = f'{priority} Priority'
                priority_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _apply_formatting(self, filepath: str):
        """Apply professional formatting to the Excel file."""
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            priority_colors = {
                'Critical': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
                'High': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
                'Medium': PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid"),
                'Low': PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            }
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Format headers
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Apply priority colors to Opportunities sheet
                if sheet_name == 'Opportunities':
                    priority_col = None
                    for col in range(1, sheet.max_column + 1):
                        if sheet.cell(row=1, column=col).value == 'Priority':
                            priority_col = col
                            break
                    
                    if priority_col:
                        for row in range(2, sheet.max_row + 1):
                            priority_value = sheet.cell(row=row, column=priority_col).value
                            if priority_value in priority_colors:
                                for col in range(1, sheet.max_column + 1):
                                    sheet.cell(row=row, column=col).fill = priority_colors[priority_value]
                
                # Apply borders to all cells
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
            
            workbook.save(filepath)
            logger.info(f"Applied formatting to Excel file: {filepath}")
            
        except Exception as e:
            logger.warning(f"Error applying Excel formatting: {e}")
    
    def _format_deadline(self, deadline) -> str:
        """Format deadline for display."""
        if not deadline:
            return 'Not specified'
        
        if isinstance(deadline, str):
            try:
                deadline = datetime.fromisoformat(deadline.replace('Z', '+00:00'))
            except:
                return deadline
        
        if isinstance(deadline, datetime):
            return deadline.strftime('%Y-%m-%d')
        
        return str(deadline)
    
    def _format_date(self, date_value) -> str:
        """Format date for display."""
        if not date_value:
            return ''
        
        if isinstance(date_value, str):
            try:
                date_value = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
            except:
                return date_value
        
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y-%m-%d %H:%M')
        
        return str(date_value)
    
    def _calculate_days_until_deadline(self, deadline) -> str:
        """Calculate days until deadline."""
        if not deadline:
            return 'Unknown'
        
        if isinstance(deadline, str):
            try:
                deadline = datetime.fromisoformat(deadline.replace('Z', '+00:00'))
            except:
                return 'Invalid date'
        
        if isinstance(deadline, datetime):
            now = datetime.now()
            if deadline.tzinfo:
                now = now.replace(tzinfo=deadline.tzinfo)
            
            days_diff = (deadline - now).days
            
            if days_diff < 0:
                return f"Expired ({abs(days_diff)} days ago)"
            elif days_diff == 0:
                return "Today"
            elif days_diff == 1:
                return "Tomorrow"
            else:
                return f"{days_diff} days"
        
        return 'Unknown'
    
    def _determine_action(self, opportunity: Dict) -> str:
        """Determine recommended action based on opportunity details."""
        priority = opportunity.get('priority', 'Low')
        deadline = opportunity.get('deadline')
        
        if priority == 'Critical':
            return 'URGENT: Review immediately'
        elif priority == 'High':
            return 'Review within 24 hours'
        elif priority == 'Medium':
            return 'Review within 3 days'
        else:
            return 'Review when convenient'

